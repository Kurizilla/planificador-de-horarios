"""
Export a schedule version to a formatted Excel file using openpyxl.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from uuid import UUID

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.db.enums import Shift
from app.db.models import (
    Project,
    ScheduleEntry,
    ScheduleVersion,
    Section,
    Subject,
    Teacher,
    TimeSlot,
)

logger = logging.getLogger(__name__)

# Day names for column headers
DAY_NAMES = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"]

# Default color palette when subject.color is not set
FALLBACK_COLORS = [
    "#3498DB", "#E67E22", "#27AE60", "#8E44AD", "#E74C3C",
    "#1ABC9C", "#D35400", "#2980B9", "#16A085", "#C0392B",
]

SHIFT_LABELS = {
    Shift.MORNING: "Matutino",
    Shift.AFTERNOON: "Vespertino",
}

# Shared styles
_thin_side = Side(style="thin")
_border = Border(
    left=_thin_side, right=_thin_side,
    top=_thin_side, bottom=_thin_side,
)
_header_font = Font(bold=True, size=11)
_title_font = Font(bold=True, size=13)
_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
_center = Alignment(horizontal="center", vertical="center")
_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _color_for_subject(subject_code: str | None, subject_color: str | None, _cache: dict[str, str] = {}) -> str:
    """Return a hex color (without '#') for a subject."""
    if subject_color:
        return subject_color.lstrip("#")
    if not subject_code:
        return "95A5A6"
    if subject_code in _cache:
        return _cache[subject_code]
    h = 0
    for ch in subject_code:
        h = ord(ch) + ((h << 5) - h)
    color = FALLBACK_COLORS[abs(h) % len(FALLBACK_COLORS)].lstrip("#")
    _cache[subject_code] = color
    return color


def _apply_header_style(cell):
    cell.font = _header_font
    cell.alignment = _center
    cell.border = _border
    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def export_schedule_to_excel(version_id: UUID, db: Session) -> bytes:
    """Generate an Excel workbook for a schedule version and return it as bytes."""

    # --- Load version and project ---
    version = db.query(ScheduleVersion).filter(ScheduleVersion.id == version_id).first()
    if not version:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Schedule version not found",
        )

    project = db.query(Project).filter(Project.id == version.project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # --- Load all entries with joins ---
    rows = (
        db.query(
            ScheduleEntry,
            Section.id.label("sec_id"),
            Section.code.label("section_code"),
            Section.grade.label("section_grade"),
            Section.shift.label("section_shift"),
            Subject.name.label("subject_name"),
            Subject.code.label("subject_code"),
            Subject.color.label("subject_color"),
            Teacher.id.label("teach_id"),
            Teacher.full_name.label("teacher_name"),
            TimeSlot.day_of_week,
            TimeSlot.start_time,
            TimeSlot.end_time,
            TimeSlot.slot_order,
        )
        .join(Section, ScheduleEntry.section_id == Section.id)
        .join(Subject, ScheduleEntry.subject_id == Subject.id)
        .join(Teacher, ScheduleEntry.teacher_id == Teacher.id)
        .join(TimeSlot, ScheduleEntry.time_slot_id == TimeSlot.id)
        .filter(ScheduleEntry.schedule_version_id == version_id)
        .order_by(TimeSlot.day_of_week, TimeSlot.slot_order)
        .all()
    )

    # --- Organize data by section ---
    # section_id -> section info
    sections_map: dict[UUID, dict] = {}
    # section_id -> { slot_key -> { day_of_week -> entry_data } }
    section_entries: dict[UUID, dict] = {}
    # Collect unique time slots
    all_slots: dict[str, dict] = {}
    # Teacher summary: teacher_id -> { name, hours, sections set }
    teacher_summary: dict[UUID, dict] = {}

    for row in rows:
        entry = row[0]
        sec_id = row.sec_id
        section_code = row.section_code
        section_grade = row.section_grade
        section_shift = row.section_shift
        subject_name = row.subject_name
        subject_code = row.subject_code
        subject_color = row.subject_color
        teach_id = row.teach_id
        teacher_name = row.teacher_name
        day_of_week = row.day_of_week
        start_time = row.start_time
        end_time = row.end_time
        slot_order = row.slot_order

        # Section info
        if sec_id not in sections_map:
            sections_map[sec_id] = {
                "code": section_code,
                "grade": section_grade,
                "shift": section_shift,
            }

        # Time slot key
        start_str = start_time.strftime("%H:%M") if hasattr(start_time, "strftime") else str(start_time)[:5]
        end_str = end_time.strftime("%H:%M") if hasattr(end_time, "strftime") else str(end_time)[:5]
        slot_key = f"{start_str}-{end_str}"
        if slot_key not in all_slots:
            all_slots[slot_key] = {"label": f"{start_str} - {end_str}", "order": slot_order}

        # Section entries
        if sec_id not in section_entries:
            section_entries[sec_id] = {}
        if slot_key not in section_entries[sec_id]:
            section_entries[sec_id][slot_key] = {}
        section_entries[sec_id][slot_key][day_of_week] = {
            "subject_name": subject_name,
            "subject_code": subject_code,
            "subject_color": subject_color,
            "teacher_name": teacher_name,
        }

        # Teacher summary
        section_label = f"G{section_grade}-{section_code}"
        if teach_id not in teacher_summary:
            teacher_summary[teach_id] = {"name": teacher_name, "hours": 0, "sections": set()}
        teacher_summary[teach_id]["hours"] += 1
        teacher_summary[teach_id]["sections"].add(section_label)

    # Sort slots by order
    sorted_slots = sorted(all_slots.items(), key=lambda x: x[1]["order"])

    # Sort sections by grade then code
    sorted_sections = sorted(
        sections_map.items(),
        key=lambda x: (x[1]["grade"], x[1]["code"]),
    )

    # --- Build workbook ---
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- Sheet per section ---
    for sec_id, sec_info in sorted_sections:
        grade = sec_info["grade"]
        code = sec_info["code"]
        shift = sec_info["shift"]
        shift_label = SHIFT_LABELS.get(shift, str(shift))
        sheet_name = f"G{grade}-{code}"
        # Excel sheet names max 31 chars
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        # Title row
        title_text = f"Horario Seccion {grade}\u00b0 {code} - {shift_label}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _title_font
        title_cell.alignment = _center

        # Column headers (row 3)
        headers = ["Hora"] + DAY_NAMES
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            _apply_header_style(cell)

        # Data rows
        entries = section_entries.get(sec_id, {})
        for row_offset, (slot_key, slot_info) in enumerate(sorted_slots):
            row_num = 4 + row_offset
            # Hora column
            hora_cell = ws.cell(row=row_num, column=1, value=slot_info["label"])
            hora_cell.alignment = _center
            hora_cell.border = _border
            hora_cell.font = Font(size=10)

            # Day columns
            for day_idx in range(5):
                col_num = 2 + day_idx
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = _border
                cell.alignment = _center_wrap

                entry_data = entries.get(slot_key, {}).get(day_idx)
                if entry_data:
                    cell.value = f"{entry_data['subject_name']}\n{entry_data['teacher_name']}"
                    color_hex = _color_for_subject(entry_data["subject_code"], entry_data["subject_color"])
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    # Use white text for readability on colored background
                    cell.font = Font(color="FFFFFF", size=10)
                else:
                    cell.value = ""

        # Column widths
        ws.column_dimensions["A"].width = 15
        for col_letter in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col_letter].width = 22

    # --- "Resumen Docentes" sheet ---
    ws_teachers = wb.create_sheet(title="Resumen Docentes")
    teacher_headers = ["Docente", "Total Horas", "Secciones"]
    for col_idx, header in enumerate(teacher_headers, 1):
        cell = ws_teachers.cell(row=1, column=col_idx, value=header)
        _apply_header_style(cell)

    sorted_teachers = sorted(teacher_summary.values(), key=lambda t: t["name"])
    for row_idx, t_info in enumerate(sorted_teachers, 2):
        ws_teachers.cell(row=row_idx, column=1, value=t_info["name"]).border = _border
        ws_teachers.cell(row=row_idx, column=1).alignment = _left_wrap

        hours_cell = ws_teachers.cell(row=row_idx, column=2, value=t_info["hours"])
        hours_cell.border = _border
        hours_cell.alignment = _center

        sections_str = ", ".join(sorted(t_info["sections"]))
        sec_cell = ws_teachers.cell(row=row_idx, column=3, value=sections_str)
        sec_cell.border = _border
        sec_cell.alignment = _left_wrap

    ws_teachers.column_dimensions["A"].width = 30
    ws_teachers.column_dimensions["B"].width = 14
    ws_teachers.column_dimensions["C"].width = 50

    # --- "Datos" sheet ---
    ws_data = wb.create_sheet(title="Datos")
    now = datetime.now(timezone.utc)

    data_rows = [
        ("Proyecto", project.name or ""),
        ("Centro Educativo", project.school_name or ""),
        ("Anio Academico", project.academic_year or ""),
        ("", ""),
        ("Version", version.version_number),
        ("Etiqueta", version.label or ""),
        ("Estado", version.status.value if version.status else ""),
        ("", ""),
        ("Generado el", now.strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("", ""),
        ("Total secciones", len(sections_map)),
        ("Total docentes", len(teacher_summary)),
        ("Total bloques asignados", len(rows)),
        ("", ""),
        ("Conflictos", version.conflicts_count or 0),
        ("Advertencias", version.warnings_count or 0),
    ]

    for row_idx, (label, value) in enumerate(data_rows, 1):
        label_cell = ws_data.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True, size=10)
        label_cell.alignment = _left_wrap

        value_cell = ws_data.cell(row=row_idx, column=2, value=value)
        value_cell.alignment = _left_wrap

    ws_data.column_dimensions["A"].width = 25
    ws_data.column_dimensions["B"].width = 40

    # --- Save to bytes ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
